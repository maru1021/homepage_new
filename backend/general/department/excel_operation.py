from sqlalchemy.orm import Session
from fastapi.responses import FileResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from .. import models
from backend.authority import models as authority_models
import os
from io import BytesIO

def export_excel(db: Session):
    departments = db.query(models.Department).all()

    if not departments:
        raise ValueError("No department data found")

    # DataFrame に変換（1列目のタイトルを「操作」に設定）
    df = pd.DataFrame([
        {"操作": "", "ID": dept.id, "部署名": dept.name}
        for dept in departments
    ])

    file_path = "departments.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())  # カラム名を追加

    for row in df.itertuples(index=False):
        ws.append(row)

    dv = DataValidation(
        type="list",
        formula1='"追加,編集,削除"',
        showDropDown=True
    )

    max_row = ws.max_row
    if max_row > 1:
        for row in range(2, max_row + 1):
            dv.add(ws[f"A{row}"])  # A列に選択肢の追加
        ws.add_data_validation(dv)

    wb.save(file_path)

    return FileResponse(file_path, filename="departments.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def import_excel(db: Session, file):
    try:
        # ExcelファイルをDataFrameに変換
        contents = file.file.read()  # 非同期でファイルを読み込む
        excel_data = BytesIO(contents)
        df = pd.read_excel(excel_data, engine="openpyxl")

        # 必須カラムの確認
        required_columns = {"操作", "ID", "部署名"}
        if not required_columns.issubset(df.columns):
            raise ValueError("Excelのフォーマットが正しくありません。'操作', 'ID', '部署名'の列が必要です。")

        for _, row in df.iterrows():
            print('row')
            print(row)
            if pd.isna(action := row["操作"]) or not action.strip():
                continue

            department_id = int(row["ID"]) if not pd.isna(row["ID"]) else None
            department_name = str(row["部署名"]).strip()

            if action.strip() == "追加":
                if department_id is not None:
                    existing_department = db.query(models.Department).filter(models.Department.id == department_id).first()
                    if existing_department:
                        raise ValueError(f"ID {department_id} は既に存在しています。")

                    existing_department = db.query(models.Department).filter(models.Department.name == department_name).first()
                    if existing_department:
                        raise ValueError(f"{department_name} は既に存在しています。")

                new_department = models.Department(name=department_name)
                db.add(new_department)
                db.commit()
                db.refresh(new_department)
                continue

            if pd.isna(row["ID"]):
                raise ValueError(f"IDを記入していない行があります。")
            if not isinstance(row["ID"], (int)) or not row["ID"]:
                raise ValueError(f"ID '{row['ID']}' を整数に修正してください。")

            elif action.strip() == "編集":
                department = db.query(models.Department).filter(models.Department.id == department_id).first()
                if not department:
                    raise ValueError(f"編集対象のID {department_id} が見つかりません。")
                department.name = department_name
                continue

            elif action.strip() == "削除":
                department = db.query(models.Department).filter(models.Department.id == department_id).first()
                if not department:
                    raise ValueError(f"削除対象のID {department_id} が見つかりません。")

                employee_count = db.query(authority_models.EmployeeAuthority.department_id).filter(
                    authority_models.EmployeeAuthority.department_id == department_id
                ).count()

                if employee_count > 0:
                    raise ValueError(f"{department.name} に所属する従業員がいるため削除できません。")

                db.delete(department)
                continue

            else:
                raise ValueError(f"無効な操作 '{action.strip()}' が含まれています。'追加', '編集', '削除' のいずれかを指定してください。")

        db.commit()
        return {"success": True, "message": "Excelデータをインポートしました"}
    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e), "field": ""}
