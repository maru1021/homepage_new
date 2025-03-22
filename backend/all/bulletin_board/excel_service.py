import io
import logging
import unicodedata
import pandas as pd
import urllib.parse
import asyncio
import base64
import hashlib
from typing import Optional, Tuple, Dict, Any, List
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import re

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload
from fastapi.responses import StreamingResponse
from fastapi import HTTPException, status

from backend.all.models import BulletinPost, BulletinCell, CellStyle, BulletinMerge
from backend.all.models import BulletinColumnDimension, BulletinRowDimension, BulletinImage

logger = logging.getLogger(__name__)

# スレッドプールエグゼキューター - CPU負荷の高い処理用
thread_executor = ThreadPoolExecutor(max_workers=4)


# ============= Excel解析関連関数 =============

import zipfile
import io
import base64
import os
import logging
import traceback


# 画像抽出関連関数
async def _extract_images_from_zip(file_data: io.BytesIO, bulletin_id: int, db: Session):
    try:
        images = []
        processed_hashes = set()  # 重複画像チェック用ハッシュセット
        logger.info(f"ZIPアプローチで画像と実際の位置・サイズ情報を抽出: bulletin_id={bulletin_id}")

        # ファイルポインタを先頭に戻す
        file_data.seek(0)

        # 列の幅と行の高さの情報を先に取得
        original_file_for_dims = io.BytesIO(file_data.read())
        file_data.seek(0)  # ポインタを戻す

        column_widths = {}
        row_heights = {}

        try:
            import openpyxl
            workbook = openpyxl.load_workbook(original_file_for_dims)
            sheet = workbook.active  # 最初のシートを対象とする

            # 列の幅を取得
            for col_letter, col_dim in sheet.column_dimensions.items():
                if col_dim.width is not None:
                    col_index = openpyxl.utils.column_index_from_string(col_letter)
                    column_widths[col_index] = col_dim.width
                    logger.info(f"列 {col_letter} (インデックス {col_index}) の幅: {col_dim.width}")

            # 行の高さを取得
            for row_index, row_dim in sheet.row_dimensions.items():
                if row_dim.height is not None:
                    row_heights[row_index] = row_dim.height
                    logger.info(f"行 {row_index} の高さ: {row_dim.height}")

            logger.info(f"シートから {len(column_widths)} 列と {len(row_heights)} 行の情報を取得しました")
        except Exception as e:
            logger.error(f"列・行情報の取得中にエラー: {str(e)}")

        # 画像情報取得用のインスタンスを作成
        try:
            # 新しいインスタンスで再度開く
            file_data.seek(0)
            workbook_for_images = openpyxl.load_workbook(file_data)

            # 実際の画像位置情報を格納するリスト
            actual_image_positions = []

            for sheet_name in workbook_for_images.sheetnames:
                sheet = workbook_for_images[sheet_name]
                if not hasattr(sheet, '_images'):
                    continue

                logger.info(f"シート '{sheet_name}' には {len(sheet._images)} 個の画像があります")

                for img_idx, img in enumerate(sheet._images):
                    try:
                        anchor = getattr(img, 'anchor', None)
                        if not anchor:
                            continue

                        # アンカーから位置情報を取得
                        from_row = getattr(anchor, 'anchorMin', None)
                        to_row = getattr(anchor, 'anchorMax', None)

                        # 異なるアンカー形式も対応
                        if not from_row and hasattr(anchor, '_from'):
                            from_row = getattr(anchor._from, 'row', 0) + 1  # 0-indexedから1-indexedへ
                            from_col = getattr(anchor._from, 'col', 0) + 1

                            if hasattr(anchor, '_to'):
                                to_row = getattr(anchor._to, 'row', from_row + 5) + 1
                                to_col = getattr(anchor._to, 'col', from_col + 5) + 1
                            else:
                                # _toがない場合は、画像サイズから推定
                                to_row = from_row + 3  # 仮の値
                                to_col = from_col + 7  # 仮の値
                        else:
                            # デフォルト値（取得できない場合）
                            from_row = 5
                            from_col = 1 + (img_idx * 9)  # 画像ごとに列をずらす
                            to_row = 8
                            to_col = from_col + 7

                        # 画像ファイル名またはpath属性
                        img_path = getattr(img, 'path', f"image{img_idx+1}.jpeg")

                        # 画像サイズ
                        width = getattr(img, 'width', 0) or 300
                        height = getattr(img, 'height', 0) or 200

                        actual_image_positions.append({
                            'path': img_path,
                            'from_row': from_row,
                            'from_col': from_col,
                            'to_row': to_row,
                            'to_col': to_col,
                            'width': width,
                            'height': height
                        })

                        logger.info(f"画像 {img_idx+1} ({img_path}) の位置: ({from_row},{from_col})-({to_row},{to_col}), サイズ: {width}x{height}")
                    except Exception as e:
                        logger.error(f"画像 {img_idx} の情報取得中にエラー: {str(e)}")

            # 位置情報の数をログ
            logger.info(f"合計 {len(actual_image_positions)} 個の画像位置情報を取得しました")
        except Exception as e:
            logger.error(f"OpenPyXLによる画像情報取得中にエラー: {str(e)}")
            actual_image_positions = []

        # ファイルポインタを先頭に戻す
        file_data.seek(0)

        # ZIPファイルとして開く
        with zipfile.ZipFile(file_data) as excel_zip:
            # ZIP内のファイル一覧
            file_list = excel_zip.namelist()
            logger.info(f"ZIP内のファイル数: {len(file_list)}")

            # 画像ファイルのパスを探す
            image_files = [name for name in file_list
                           if name.startswith('xl/media/') and
                              (name.lower().endswith('.jpg') or name.lower().endswith('.jpeg'))]

            logger.info(f"見つかったJPEG画像ファイル: {len(image_files)}")

            # 各画像ファイルを処理
            for idx, img_path in enumerate(image_files):
                try:
                    # 画像データを取得
                    img_data = excel_zip.read(img_path)

                    # 画像データのハッシュを計算して重複チェック
                    img_hash = hashlib.md5(img_data).hexdigest()
                    if img_hash in processed_hashes:
                        logger.info(f"重複画像をスキップ: {img_path}")
                        continue

                    processed_hashes.add(img_hash)

                    # ヘッダーチェック (JPEGシグネチャ)
                    if img_data[:2] != b'\xff\xd8':
                        logger.info(f"JPEG以外の画像をスキップ: {img_path}")
                        continue

                    # 画像のBase64エンコード
                    image_b64 = base64.b64encode(img_data).decode('utf-8')

                    # 位置情報を取得 - 実際の画像位置情報から検索
                    pos_info = None
                    # 完全一致またはファイル名部分一致で検索
                    file_name = img_path.split('/')[-1]
                    for pos in actual_image_positions:
                        if pos['path'] == img_path or file_name in pos['path'] or f"image{idx+1}" in pos['path']:
                            pos_info = pos
                            break

                    if pos_info:
                        # 取得した位置情報を使用
                        from_row = pos_info['from_row']
                        from_col = pos_info['from_col']
                        to_row = pos_info['to_row']
                        to_col = pos_info['to_col']
                        width = pos_info['width']
                        height = pos_info['height']

                        logger.info(f"画像 {img_path} は実際の位置情報を使用: ({from_row},{from_col})-({to_row},{to_col})")
                    else:
                        # 位置情報がない場合はインデックスから推定
                        # 画像が左右に並ぶように配置
                        if idx == 0:
                            from_row = 5
                            from_col = 1
                            to_row = 8
                            to_col = 8
                        elif idx == 1:
                            from_row = 5
                            from_col = 9
                            to_row = 8
                            to_col = 16
                        else:
                            from_row = 5 + (idx // 2) * 4  # 行方向にも展開
                            from_col = 1 + (idx % 2) * 9   # 2列で交互に配置
                            to_row = from_row + 3
                            to_col = from_col + 7

                        width = 300
                        height = 200

                        logger.info(f"画像 {img_path} は推定位置を使用: ({from_row},{from_col})-({to_row},{to_col})")

                    # セルの幅と高さから画像の実際のピクセルサイズを計算
                    pixel_width = 0
                    for col in range(from_col, to_col + 1):
                        cell_width = column_widths.get(col, 8.43)  # デフォルト幅
                        pixel_width += cell_width * 9  # ピクセル変換係数

                    pixel_height = 0
                    for row in range(from_row, to_row + 1):
                        cell_height = row_heights.get(row, 20)  # デフォルト高さ
                        pixel_height += cell_height

                    # 計算したピクセルサイズをログ
                    logger.info(f"画像 {img_path} のセルから計算したサイズ: {pixel_width}x{pixel_height}px")

                    # セルから計算したサイズを使用するが、最小値は設定
                    if pixel_width < 50:
                        pixel_width = 300
                    if pixel_height < 50:
                        pixel_height = 200

                    # 画像オブジェクトを作成
                    bulletin_image = BulletinImage(
                        bulletin_id=bulletin_id,
                        image_data=image_b64,
                        image_type="jpeg",
                        from_row=from_row,
                        from_col=from_col,
                        to_row=to_row,
                        to_col=to_col,
                        width=pixel_width,  # セルから計算したピクセル幅
                        height=pixel_height  # セルから計算したピクセル高さ
                    )

                    images.append(bulletin_image)
                except Exception as e:
                    logger.error(f"画像 {img_path} の処理中にエラー: {str(e)}")
                    logger.error(traceback.format_exc())

            # 画像データの一括保存
            if images:
                db.bulk_save_objects(images)
                logger.info(f"保存された画像の数: {len(images)}")
            else:
                logger.warning("JPEG画像が見つかりませんでした")

        return images

    except Exception as e:
        logger.error(f"画像抽出中のエラー: {str(e)}")
        logger.error(traceback.format_exc())
        return []

# Excelファイルを解析してデータベースに保存する関数
async def parse_excel_to_db(
    file_data: io.BytesIO, filename: str, employee_id: int,
    title: str, content: Optional[str], db: Session
) -> BulletinPost:
    try:
        # 掲示板投稿を作成
        bulletin_post = BulletinPost(
            title=title, content=content,
            employee_id=employee_id, filename=filename
        )
        db.add(bulletin_post)
        db.flush()

        # 従来の方法を試す（DFとOpenpyxlでの処理）
        df_task = asyncio.create_task(_run_in_thread(
            lambda: pd.read_excel(file_data, engine='openpyxl')
        ))

        # ファイルポインタを先頭に戻す
        file_data.seek(0)

        workbook_task = asyncio.create_task(_run_in_thread(
            lambda: openpyxl.load_workbook(file_data)
        ))

        # 両方のタスクが完了するのを待つ
        df = await df_task
        workbook = await workbook_task

        # データ処理
        cell_id_mapping = await _process_cell_data(df, bulletin_post.id, db)
        await _process_styles_and_properties(workbook.active, cell_id_mapping, bulletin_post.id, db)

        # ZIP方式で直接抽出 (JPEG画像のみを抽出)
        file_data.seek(0)
        images = await _extract_images_from_zip(file_data, bulletin_post.id, db)

        db.commit()
        return bulletin_post

    except Exception as e:
        db.rollback()
        error_message = f"エクセル解析エラー: {filename} - {str(e)}"
        logger.error(error_message)
        raise Exception(error_message)


# セルデータを処理しマッピングを返す
async def _process_cell_data(df, bulletin_id, db):
    # 一括でセルオブジェクトを作成して処理を効率化
    cells_to_add = [
        BulletinCell(
            bulletin_id=bulletin_id,
            row=row_idx + 1,  # pandasは0始まりなので+1
            col=col_idx,
            value=str(value) if pd.notna(value) else None
        )
        for row_idx, row in df.iterrows()
        for col_idx, value in enumerate(row, 1)
    ]

    # セルを一括で追加
    db.bulk_save_objects(cells_to_add)
    db.flush()

    # ID取得のためにもう一度クエリ実行（bulk_save_objectsでは自動生成IDが取得できないため）
    cells = db.query(BulletinCell).filter(BulletinCell.bulletin_id == bulletin_id).all()
    return {(cell.row, cell.col): cell.id for cell in cells}


# スタイルとシート特性を処理
async def _process_styles_and_properties(sheet, cell_id_mapping, bulletin_id, db):
    # データ収集タスクを並行実行
    style_task = asyncio.create_task(_collect_cell_styles(sheet, cell_id_mapping))
    merge_task = asyncio.create_task(_collect_merge_ranges(sheet, bulletin_id))
    col_dim_task = asyncio.create_task(_collect_column_dimensions(sheet, bulletin_id))
    row_dim_task = asyncio.create_task(_collect_row_dimensions(sheet, bulletin_id))

    # 全てのタスクから結果を取得
    styles, merges, col_dims, row_dims = await asyncio.gather(
        style_task, merge_task, col_dim_task, row_dim_task
    )

    # 一括でデータベースに保存
    for data_list in [styles, merges, col_dims, row_dims]:
        if data_list:
            db.bulk_save_objects(data_list)


# シートからセルスタイル情報を収集
async def _collect_cell_styles(sheet, cell_id_mapping):
    def process_cells():
        return [
            CellStyle(cell_id=cell_id_mapping[(row_idx, col_idx)], **cell_style)
            for row_idx, row in enumerate(sheet.iter_rows(), 1)
            for col_idx, cell in enumerate(row, 1)
            if (cell_style := _parse_cell_style(cell)) and (row_idx, col_idx) in cell_id_mapping
        ]

    # スレッドプールで実行
    return await _run_in_thread(process_cells)


# シートから結合セル情報を収集
async def _collect_merge_ranges(sheet, bulletin_id):
    return [
        BulletinMerge(
            bulletin_id=bulletin_id,
            start_row=merged_range.min_row,
            start_col=merged_range.min_col,
            end_row=merged_range.max_row,
            end_col=merged_range.max_col
        )
        for merged_range in sheet.merged_cells.ranges
    ]


# シートから列幅情報を収集
async def _collect_column_dimensions(sheet, bulletin_id):
    return [
        BulletinColumnDimension(
            bulletin_id=bulletin_id,
            col=openpyxl.utils.column_index_from_string(col_letter),
            width=col_dim.width
        )
        for col_letter, col_dim in sheet.column_dimensions.items()
        if col_dim.width is not None
    ]


# シートから行高情報を収集
async def _collect_row_dimensions(sheet, bulletin_id):
    return [
        BulletinRowDimension(
            bulletin_id=bulletin_id,
            row=row_idx,
            height=row_dim.height
        )
        for row_idx, row_dim in sheet.row_dimensions.items()
        if row_dim.height is not None
    ]


# 既存の掲示板投稿のExcelデータを更新する
async def update_excel_in_db(
    bulletin_id: int, file_data: io.BytesIO,
    filename: str, title: str, content: Optional[str], db: Session
) -> BulletinPost:
    try:
        # 関連データ削除
        await _delete_related_data(bulletin_id, db)

        # 基本情報更新
        post = db.query(BulletinPost).filter(BulletinPost.id == bulletin_id).first()
        post.title = title
        post.content = content
        post.filename = filename
        post.updated_at = datetime.utcnow()
        db.flush()

        # 並行処理でデータ読み込み
        df_task = asyncio.create_task(_run_in_thread(
            lambda: pd.read_excel(file_data, engine='openpyxl')
        ))

        file_data.seek(0)

        workbook_task = asyncio.create_task(_run_in_thread(
            lambda: openpyxl.load_workbook(file_data, data_only=True)
        ))

        df, workbook = await asyncio.gather(df_task, workbook_task)
        sheet = workbook.active

        # データ処理
        cell_id_mapping = await _process_cell_data(df, bulletin_id, db)
        await _process_styles_and_properties(sheet, cell_id_mapping, bulletin_id, db)

        # ZIP方式でJPEG画像のみを抽出
        file_data.seek(0)
        await _extract_images_from_zip(file_data, bulletin_id, db)

        db.commit()
        return post

    except Exception as e:
        db.rollback()
        error_message = f"エクセル更新エラー: ID {bulletin_id} - {str(e)}"
        logger.error(error_message)
        raise Exception(error_message)


# 掲示板投稿の関連データを削除
async def _delete_related_data(bulletin_id, db):
    from sqlalchemy import select
    cell_ids_query = select(BulletinCell.id).where(BulletinCell.bulletin_id == bulletin_id)

    # 関連データを一括削除
    db.query(CellStyle).filter(CellStyle.cell_id.in_(cell_ids_query)).delete(synchronize_session=False)
    db.query(BulletinCell).filter(BulletinCell.bulletin_id == bulletin_id).delete(synchronize_session=False)
    db.query(BulletinMerge).filter(BulletinMerge.bulletin_id == bulletin_id).delete(synchronize_session=False)
    db.query(BulletinColumnDimension).filter(BulletinColumnDimension.bulletin_id == bulletin_id).delete(synchronize_session=False)
    db.query(BulletinRowDimension).filter(BulletinRowDimension.bulletin_id == bulletin_id).delete(synchronize_session=False)
    db.query(BulletinImage).filter(BulletinImage.bulletin_id == bulletin_id).delete(synchronize_session=False)


# ============= Excel生成関連関数 =============

# データベースのデータからExcelファイルを生成する
async def generate_excel_from_db(bulletin_id: int, db: Session) -> Tuple[io.BytesIO, str]:
    # 掲示板投稿と関連データ取得
    post = db.query(BulletinPost).filter(BulletinPost.id == bulletin_id).first()
    if not post:
        raise Exception(f"ID {bulletin_id} の掲示板投稿が見つかりません")

    # 並行して各種データを取得
    cells_task = _run_in_thread(
        lambda: db.query(BulletinCell).filter(BulletinCell.bulletin_id == bulletin_id).all()
    )
    merges_task = _run_in_thread(
        lambda: db.query(BulletinMerge).filter(BulletinMerge.bulletin_id == bulletin_id).all()
    )
    column_dims_task = _run_in_thread(
        lambda: db.query(BulletinColumnDimension).filter(BulletinColumnDimension.bulletin_id == bulletin_id).all()
    )
    row_dims_task = _run_in_thread(
        lambda: db.query(BulletinRowDimension).filter(BulletinRowDimension.bulletin_id == bulletin_id).all()
    )

    # 各タスクの完了を待機
    cells, merges, column_dims, row_dims = await asyncio.gather(
        cells_task, merges_task, column_dims_task, row_dims_task
    )

    # セルスタイル情報を効率的に取得
    cell_ids = [cell.id for cell in cells]
    styles = db.query(CellStyle).filter(CellStyle.cell_id.in_(cell_ids)).all() if cell_ids else []
    style_map = {style.cell_id: style for style in styles}

    # Excelファイル作成処理をスレッドプールで実行
    output = await _run_in_thread(
        lambda: _create_excel_file(cells, merges, column_dims, row_dims, style_map)
    )

    return output, _generate_safe_filename(post)


# Excelファイルを作成する
def _create_excel_file(cells, merges, column_dims, row_dims, style_map):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # セルデータとスタイル適用
    for cell_data in cells:
        sheet_cell = sheet.cell(row=cell_data.row, column=cell_data.col, value=cell_data.value)
        if cell_data.id in style_map:
            _apply_cell_style(sheet_cell, style_map[cell_data.id])

    # セル結合適用
    for merge in merges:
        sheet.merge_cells(
            start_row=merge.start_row, start_column=merge.start_col,
            end_row=merge.end_row, end_column=merge.end_col
        )

    # 列幅と行高設定
    for col_dim in column_dims:
        sheet.column_dimensions[get_column_letter(col_dim.col)].width = col_dim.width

    for row_dim in row_dims:
        sheet.row_dimensions[row_dim.row].height = row_dim.height

    # ファイル出力
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


# ============= レスポンス関連関数 =============

# 掲示板投稿をレスポンス形式に整形する
async def format_bulletin_response(post: BulletinPost, employee_name: str, db: Session) -> Dict[str, Any]:
    return {
        "id": post.id,
        "title": post.title,
        "content": post.content,
        "employee_id": post.employee_id,
        "employee_name": employee_name,
        "created_at": post.created_at,
        "updated_at": post.updated_at,
        "filename": post.filename
    }


# 掲示板投稿一覧を取得する
async def get_bulletin_list(skip: int, limit: int, db: Session) -> Dict[str, Any]:
    # 並行してデータ取得
    count_task = asyncio.create_task(_run_in_thread(
        lambda: db.query(BulletinPost).count()
    ))

    # 投稿データをjoinedloadを使って従業員情報と画像情報を一緒に取得
    posts_task = asyncio.create_task(_run_in_thread(
        lambda: db.query(BulletinPost)
               .options(
                   joinedload(BulletinPost.employee),
                   joinedload(BulletinPost.images)
               )
               .order_by(BulletinPost.created_at.desc())
               .offset(skip)
               .limit(limit)
               .all()
    ))

    # データ取得を待機
    total_count, posts = await asyncio.gather(count_task, posts_task)

    # レスポンスデータを作成
    posts_data = [
        {
            "id": post.id,
            "title": post.title,
            "content": post.content,
            "employee_id": post.employee_id,
            "employee_name": post.employee.name if post.employee else None,
            "created_at": post.created_at,
            "updated_at": post.updated_at,
            "filename": post.filename,
            "images": [
                {
                    "image_data": image.image_data,
                    "image_type": image.image_type,
                    "from_row": image.from_row,
                    "from_col": image.from_col,
                    "to_row": image.to_row,
                    "to_col": image.to_col,
                    "width": image.width,
                    "height": image.height
                }
                for image in post.images
            ] if post.images else []
        }
        for post in posts
    ]

    return {
        "posts": posts_data,
        "total": total_count
    }

# エクセルファイルをダウンロードするためのレスポンスを作成する
async def create_excel_response(output: io.BytesIO, filename: str) -> StreamingResponse:
    # ファイル名をURLエンコード
    encoded_filename = urllib.parse.quote(filename)

    # レスポンスの設定
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# 掲示板投稿の詳細情報を取得する
async def get_bulletin_detail(bulletin_id: int, db: Session) -> Dict[str, Any]:
    # 投稿の取得（employeeとjoinedloadで一緒に取得）
    post = db.query(BulletinPost).options(joinedload(BulletinPost.employee)).filter(BulletinPost.id == bulletin_id).first()
    if not post:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"ID {bulletin_id} の掲示板投稿が見つかりません")

    # 並行して各データを取得
    cells_data, merges_data, column_dimensions_data, row_dimensions_data, images_data = await asyncio.gather(
        _fetch_bulletin_cells_with_styles(db, bulletin_id),
        _fetch_bulletin_merges(db, bulletin_id),
        _fetch_bulletin_column_dimensions(db, bulletin_id),
        _fetch_bulletin_row_dimensions(db, bulletin_id),
        _fetch_bulletin_images(db, bulletin_id)  # 画像データの取得を追加
    )

    employee_name = post.employee.name if post.employee else None

    # レスポンスデータを作成
    return {
        "id": post.id,
        "title": post.title,
        "content": post.content,
        "employee_id": post.employee_id,
        "employee_name": employee_name,
        "created_at": post.created_at,
        "updated_at": post.updated_at,
        "filename": post.filename,
        "cells": cells_data,
        "merges": merges_data,
        "column_dimensions": column_dimensions_data,
        "row_dimensions": row_dimensions_data,
        "images": images_data,  # 画像データを追加
        "exists": True,
        "parsed_at": datetime.now().isoformat()
    }


# ============= データ取得関連関数 =============

# セルデータとスタイル情報を効率的に取得
async def _fetch_bulletin_cells_with_styles(db: Session, bulletin_id: int) -> List[Dict[str, Any]]:
    # セルデータとスタイル情報を並行して取得
    cells = await _run_in_thread(
        lambda: db.query(BulletinCell).filter(BulletinCell.bulletin_id == bulletin_id).all()
    )

    cell_ids = [cell.id for cell in cells]

    if not cell_ids:
        return []

    # joinedloadを使用してセルとスタイルを一緒に取得
    cells_with_styles = await _run_in_thread(
        lambda: db.query(BulletinCell).options(
            joinedload(BulletinCell.style)
        ).filter(
            BulletinCell.id.in_(cell_ids)
        ).all()
    )

    # セルデータを整形
    return [_format_cell_data(cell, cell.style) for cell in cells_with_styles]


# 結合セル情報を取得
async def _fetch_bulletin_merges(db: Session, bulletin_id: int) -> List[Dict[str, Any]]:
    merges = await _run_in_thread(
        lambda: db.query(BulletinMerge).filter(BulletinMerge.bulletin_id == bulletin_id).all()
    )

    return [
        {
            "start": {"row": merge.start_row, "col": merge.start_col},
            "end": {"row": merge.end_row, "col": merge.end_col}
        }
        for merge in merges
    ]


# 列の幅情報を取得
async def _fetch_bulletin_column_dimensions(db: Session, bulletin_id: int) -> Dict[str, float]:
    column_dimensions = await _run_in_thread(
        lambda: db.query(BulletinColumnDimension).filter(
            BulletinColumnDimension.bulletin_id == bulletin_id
        ).all()
    )

    return {str(col_dim.col): col_dim.width for col_dim in column_dimensions}


# 行の高さ情報を取得
async def _fetch_bulletin_row_dimensions(db: Session, bulletin_id: int) -> Dict[str, float]:
    row_dimensions = await _run_in_thread(
        lambda: db.query(BulletinRowDimension).filter(
            BulletinRowDimension.bulletin_id == bulletin_id
        ).all()
    )

    return {str(row_dim.row): row_dim.height for row_dim in row_dimensions}


# 画像情報を取得
async def _fetch_bulletin_images(db: Session, bulletin_id: int) -> List[Dict[str, Any]]:
    images = await _run_in_thread(
        lambda: db.query(BulletinImage).filter(
            BulletinImage.bulletin_id == bulletin_id
        ).all()
    )

    return [
        {
            "image_data": image.image_data,
            "image_type": image.image_type,
            "from_row": image.from_row,
            "from_col": image.from_col,
            "to_row": image.to_row,
            "to_col": image.to_col,
            "width": image.width,
            "height": image.height
        }
        for image in images
    ]


# ============= ユーティリティ関数 =============

# 色情報をaRGB形式に修正する
def _fix_color_format(color_str):
    if not color_str or not isinstance(color_str, str):
        return None

    # 透明色または空の色
    if color_str == "00000000" or color_str == "":
        return None

    # '#'で始まる場合は削除
    if color_str.startswith('#'):
        color_str = color_str[1:]

    # 形式に応じた処理
    if len(color_str) == 8:
        return color_str
    elif len(color_str) == 6:
        return "FF" + color_str

    # その他は黒色（不透明）
    return "FF000000"


# セルのスタイル情報を解析
def _parse_cell_style(cell) -> Dict[str, Any]:
    style = {}

    # フォント情報
    if cell.font:
        style['font_bold'] = cell.font.bold

        try:
            style['font_color'] = str(cell.font.color.rgb)[:64] if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb else None
        except Exception:
            style['font_color'] = None

        try:
            style['font_size'] = float(cell.font.size) if cell.font.size else None
        except (TypeError, ValueError):
            style['font_size'] = None
    else:
        style['font_bold'] = False
        style['font_color'] = None
        style['font_size'] = None

    # 背景色
    try:
        style['bg_color'] = str(cell.fill.fgColor.rgb)[:64] if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor.rgb else None
    except Exception:
        style['bg_color'] = None

    # 罫線スタイル
    if cell.border:
        for side in ['top', 'right', 'bottom', 'left']:
            border_side = getattr(cell.border, side, None)
            if border_side:
                style[f'border_{side}_style'] = str(border_side.style)[:16] if border_side.style else None
                try:
                    style[f'border_{side}_color'] = str(border_side.color.rgb)[:64] if hasattr(border_side.color, 'rgb') and border_side.color.rgb else None
                except Exception:
                    style[f'border_{side}_color'] = None

    # 配置情報
    if cell.alignment:
        style['alignment_horizontal'] = str(cell.alignment.horizontal)[:16] if cell.alignment.horizontal else None
        style['alignment_vertical'] = str(cell.alignment.vertical)[:16] if cell.alignment.vertical else None

    return style


# セルにスタイルを適用
def _apply_cell_style(cell, style):
    # フォントスタイル設定
    if style.font_bold or style.font_color or style.font_size:
        font_args = {'bold': style.font_bold, 'size': style.font_size}
        if style.font_color:
            fixed_color = _fix_color_format(style.font_color)
            if fixed_color:
                font_args['color'] = fixed_color
        cell.font = Font(**font_args)

    # 背景色設定
    if style.bg_color:
        fixed_bg_color = _fix_color_format(style.bg_color)
        if fixed_bg_color:
            cell.fill = PatternFill(fill_type='solid', fgColor=fixed_bg_color)

    # 罫線設定
    borders = {}
    for side in ['top', 'right', 'bottom', 'left']:
        style_attr = getattr(style, f'border_{side}_style', None)
        if style_attr:
            color_attr = getattr(style, f'border_{side}_color', None)
            color = _fix_color_format(color_attr) if color_attr else None
            borders[side] = Side(style=style_attr, color=color)

    if borders:
        cell.border = Border(**borders)

    # 配置設定
    if style.alignment_horizontal or style.alignment_vertical:
        cell.alignment = Alignment(
            horizontal=style.alignment_horizontal,
            vertical=style.alignment_vertical
        )


# 安全なファイル名を生成
def _generate_safe_filename(post):
    if post.filename:
        # 拡張子保持
        name_part, ext_part = post.filename.rsplit('.', 1) if '.' in post.filename else (post.filename, 'xlsx')
        # ASCII化と特殊文字処理
        ascii_name = unicodedata.normalize('NFKD', name_part).encode('ASCII', 'ignore').decode()
        safe_name = ''.join(c if c.isalnum() or c in '_- ' else '_' for c in ascii_name)
        return f"{safe_name}.{ext_part}"
    else:
        # タイトルからファイル名生成
        ascii_title = unicodedata.normalize('NFKD', post.title).encode('ASCII', 'ignore').decode()
        safe_title = ''.join(c if c.isalnum() or c in '_- ' else '_' for c in ascii_title)
        return f"{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# セルデータをフロントエンド用にフォーマット
def _format_cell_data(cell, style):
    cell_data = {
        "row": cell.row,
        "col": cell.col,
        "value": cell.value
    }

    # スタイル情報があれば追加
    if style:
        cell_data["style"] = {
            "font": {
                "bold": style.font_bold,
                "color": style.font_color,
                "size": style.font_size
            },
            "fill": {
                "bgColor": style.bg_color
            },
            "border": {
                "top": {
                    "style": style.border_top_style,
                    "color": style.border_top_color
                },
                "right": {
                    "style": style.border_right_style,
                    "color": style.border_right_color
                },
                "bottom": {
                    "style": style.border_bottom_style,
                    "color": style.border_bottom_color
                },
                "left": {
                    "style": style.border_left_style,
                    "color": style.border_left_color
                }
            },
            "alignment": {
                "horizontal": style.alignment_horizontal,
                "vertical": style.alignment_vertical
            }
        }

    return cell_data


# スレッドプールエグゼキューターでCPU負荷の高い処理を実行
async def _run_in_thread(func):
    return await asyncio.get_event_loop().run_in_executor(thread_executor, func)